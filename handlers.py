from aiogram import Dispatcher, html, Router, Bot
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, FSInputFile, CallbackQuery
from openpyxl.reader.excel import load_workbook
from sqlalchemy import not_
from sqlalchemy.orm import Session
from datetime import datetime, date, time, timedelta
from database import SessionLocal, Employee, Attendance
from keyboards import get_reply_keyboard, create_date_keyboard, get_reply_type_keyboard, get_supervisor_keyboard, get_return_keyboard
import pandas as pd
import os
from aiogram.fsm.state import State, StatesGroup
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
from utils.months import months_russian
from utils.colors import *


class Form(StatesGroup):
    waiting_for_full_name = State()
    choosing_departure_type = State()
    choosing_supervisor = State()
    waiting_for_reason = State()
    waiting_for_departure_time = State()


async def send_report(message: Message, selected_date: str) -> None:
    db: Session = SessionLocal()
    #employees = db.query(Employee).all()
    PRODUCT = os.getenv('PRODUCT')

    if int(PRODUCT):
        IGNORE_WORKERS = ['1195996440', '6468224924']

    else:
        IGNORE_WORKERS = []

    employees = db.query(Employee).filter(not_(Employee.telegram_id.in_(IGNORE_WORKERS))).all()

    data = []
    for employee in employees:
        attendance = db.query(Attendance).filter(Attendance.employee_id == employee.id,
                                                 Attendance.date == selected_date).first()

        data.append({
            "ФИО": employee.fio,
            "Дата": selected_date,
            "Телеграмм - ID": employee.telegram_id,
            "Телеграмм Ник": employee.full_name,
            "Время прибытия": attendance.arrival_time if attendance else '',
            "Время ухода": attendance.departure_time if attendance else '',
            "Тип отъезда": attendance.departure_type if attendance else '',
            "Руководитель": attendance.supervisor if attendance else '',
            "Причина": attendance.departure_reason if attendance else '',
            "Время отъезда": attendance.departure_time_actual if attendance else '',
            "Время возвращения": attendance.return_time if attendance else '',
        })

    df = pd.DataFrame(data)
    report_file = f"Отчёт({selected_date}).xlsx"
    df.to_excel(report_file, index=False, engine='openpyxl')

    wb = load_workbook(report_file)
    ws = wb.active
    ws.title = "Отчет"

    ws.insert_rows(1)
    ws.merge_cells('A1:K1')
    ws['A1'] = f'Отчет сотрудников за {selected_date}'
    ws['A1'].font = Font(size=16, bold=True)

    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        if column_letter:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    for row in ws.iter_rows(min_row=1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = black_border

    wb.save(report_file)

    report_document = FSInputFile(report_file)
    await message.answer_document(report_document, caption=f"Отчет сотрудников за {selected_date}")

    db.close()

date_router = Router()


@date_router.callback_query(lambda c: c.data and c.data.startswith("report_"))
async def process_date_callback(callback_query: CallbackQuery):
    selected_date = callback_query.data.split("_")[1]
    await send_report(callback_query.message, selected_date)
    await callback_query.answer()

def register_handlers(dp: Dispatcher) -> None:
    @dp.message(Command(commands=['start']))
    async def command_start_handler(message: Message, state: FSMContext) -> None:

        telegram_id = message.from_user.id
        full_name = message.from_user.full_name

        db: Session = SessionLocal()

        employee = db.query(Employee).filter(Employee.telegram_id == telegram_id).first()

        if not employee:
            welcome_text = (
                f"Здравствуйте, {html.bold(message.from_user.full_name)}! Добро пожаловать в систему учета времени."
            )
            await message.answer(welcome_text)

            await message.answer("Пожалуйста, укажите ваше ФИО.")
            await state.set_state(Form.waiting_for_full_name)

        else:
            ADMIN_ID = os.getenv('ADMIN_ID')
            is_admin = str(message.from_user.id) in ADMIN_ID
            keyboard = get_reply_keyboard(is_admin)
            await message.answer(
                f"Здравствуйте, {html.bold(full_name)}! Вы уже зарегистрированы.",
                reply_markup=keyboard
            )

        db.close()

    @dp.message(Form.waiting_for_full_name)
    async def handle_full_name(message: Message, state: FSMContext) -> None:
        telegram_id = message.from_user.id
        full_name = message.from_user.full_name

        db = SessionLocal()
        try:
            existing_employee = db.query(Employee).filter(Employee.telegram_id == telegram_id).first()
            if existing_employee:
                existing_employee.fio = message.text
                db.commit()
                db.refresh(existing_employee)
            else:
                new_employee = Employee(
                    telegram_id=telegram_id,
                    full_name=full_name,
                    fio=message.text
                )
                db.add(new_employee)
                db.commit()
                db.refresh(new_employee)

            ADMIN_ID = os.getenv('ADMIN_ID')
            is_admin = str(message.from_user.id) in ADMIN_ID
            keyboard = get_reply_keyboard(is_admin)
            await message.answer(
                f"Здравствуйте, {html.bold(full_name)}! Вы успешно зарегистрированы.",
                reply_markup=keyboard
            )
            await state.clear()  #
        except Exception as e:
            await message.answer(f"Произошла ошибка: {str(e)}")
        finally:
            db.close()

    @date_router.message(lambda message: message.text == "Пришел")
    async def arrival_handler(message: Message) -> None:
        telegram_id = message.from_user.id
        current_date = date.today()
        current_time = datetime.now().time()
        late_time = time(9, 5)

        db: Session = SessionLocal()
        ADMIN_ID = os.getenv('ADMIN_ID')
        is_admin = str(message.from_user.id) in ADMIN_ID
        keyboard = get_reply_keyboard(is_admin)

        # Найдите сотрудника по Telegram ID
        employee = db.query(Employee).filter(Employee.telegram_id == telegram_id).first()
        if employee:
            # Найдите запись о присутствии на сегодня
            attendance = db.query(Attendance).filter(
                Attendance.employee_id == employee.id,
                Attendance.date == current_date
            ).first()

            if attendance:
                if attendance.check:
                    # Если запись есть и она еще не проверена, обновите её
                    attendance.arrival_time = current_time.replace(microsecond=0)
                    attendance.check = False
                    db.commit()
                    await message.answer("Время прибытия успешно отмечено!", reply_markup=keyboard)
                else:
                    # Если запись уже была проверена
                    await message.answer("Вы уже отметили прибытие сегодня.", reply_markup=keyboard)
            else:
                # Если записи о присутствии нет, создайте её
                is_late = current_time > late_time
                new_attendance = Attendance(
                    employee_id=employee.id,
                    date=current_date,
                    arrival_time=current_time.replace(microsecond=0),
                    late=is_late
                )
                db.add(new_attendance)
                db.commit()
                await message.answer(
                    "Время прибытия успешно отмечено!" if not is_late
                    else "Время прибытия успешно отмечено! К сожалению, вы опоздали."
                )
        else:
            # Если сотрудник не зарегистрирован
            await message.answer("Вы не зарегистрированы. Пожалуйста, используйте команду /start для регистрации.",
                                 reply_markup=keyboard)

        db.close()

    @date_router.message(lambda message: message.text == "Ушел")
    async def departure_handler(message: Message) -> None:
        telegram_id = message.from_user.id
        current_date = date.today()
        current_time = datetime.now().time()
        ADMIN_ID = os.getenv('ADMIN_ID')
        is_admin = str(message.from_user.id) in ADMIN_ID
        keyboard = get_reply_keyboard(is_admin)
        db: Session = SessionLocal()

        employee = db.query(Employee).filter(Employee.telegram_id == telegram_id).first()
        if employee:
            attendance = db.query(Attendance).filter(Attendance.employee_id == employee.id,
                                                     Attendance.date == current_date).first()
            if attendance and attendance.arrival_time:
                attendance.departure_time = current_time.replace(microsecond=0)
                db.commit()
                await message.answer("Время ухода успешно отмечено!", reply_markup=keyboard)
            else:
                await message.answer("Вы не отметили прибытие или уже отметили уход.", reply_markup=keyboard)
        else:
            await message.answer("Вы не зарегистрированы. Пожалуйста, используйте команду /start для регистрации.", reply_markup=keyboard)

        db.close()

    @dp.message(Command(commands=['report']))
    async def report_handler(message: Message) -> None:
        db: Session = SessionLocal()

        command_parts = message.text.split()
        if len(command_parts) > 1:
            date_str = command_parts[1]
            try:
                report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                await message.answer("Неверный формат даты. Используйте формат YYYY-MM-DD.")
                db.close()
                return
        else:
            report_date = date.today()

        IGNORE_WORKERS = ['1195996440', '6468224924']
        employees = db.query(Employee).filter(not_(Employee.telegram_id.in_(IGNORE_WORKERS))).all()

        attendances = db.query(Attendance).filter(Attendance.date == report_date).all()

        attendances_dict = {}
        for attendance in attendances:
            attendances_dict[attendance.employee_id] = {
                "Время прибытия": attendance.arrival_time,
                "Время ухода": attendance.departure_time,
                "Тип отъезда": attendance.departure_type,
                "Руководитель": attendance.supervisor,
                "Причина": attendance.departure_reason,
                "Время отъезда": attendance.departure_time_actual,
                "Время возвращения": attendance.return_time
            }

        data = []
        for employee in employees:
            attendance = attendances_dict.get(employee.id, {
                'Время прибытия': '',
                'Время ухода': '',
                'Тип отъезда': '',
                'Руководитель': '',
                'Причина': '',
                'Время отъезда': '',
                'Время возвращения': ''
            })
            data.append({
                "ФИО": employee.fio,
                "Дата": report_date,
                "Телеграмм - ID": employee.telegram_id,
                "Телеграмм Ник": employee.full_name,
                "Время прибытия": attendance['Время прибытия'],
                "Время ухода": attendance['Время ухода'],
                "Тип отъезда": attendance['Тип отъезда'],
                "Руководитель": attendance['Руководитель'],
                "Причина": attendance['Причина'],
                "Время отъезда": attendance['Время отъезда'],
                "Время возвращения": attendance['Время возвращения'],
            })

        df = pd.DataFrame(data)
        report_file = f"Отчёт({report_date}).xlsx"
        df.to_excel(report_file, index=False, engine='openpyxl')

        wb = load_workbook(report_file)
        ws = wb.active
        ws.title = "Отчет"

        ws.insert_rows(1)
        ws.merge_cells('A1:K1')
        ws['A1'] = f'Отчет сотрудников за {report_date}'
        ws['A1'].font = Font(size=16, bold=True)

        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            if column_letter:
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.iter_rows(min_row=1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = black_border

        wb.save(report_file)

        report_document = FSInputFile(report_file)
        await message.answer_document(report_document, caption=f"Отчет сотрудников за {report_date}")

        db.close()

    @date_router.message(lambda message: message.text == "Отчет")
    async def report_button_handler(message: Message) -> None:
        ADMIN_ID = os.getenv('ADMIN_ID')
        if str(message.from_user.id) not in ADMIN_ID:
            await message.answer("У вас нет прав для выполнения этого действия.")
            return
        keyboard = create_date_keyboard()
        await message.answer(
            "Чтобы получить отчет за конкретный день, напишите команду в формате: /report YYYY-MM-DD, где YYYY-MM-DD — это дата в формате год-месяц-день. Например: /report 2024-07-20.")
        await message.answer("Выберите дату для отчета:", reply_markup=keyboard)

    @date_router.message(lambda message: message.text == "Опоздуны")
    async def late_report_handler(message: Message) -> None:
        ADMIN_ID = os.getenv('ADMIN_ID')
        if str(message.from_user.id) not in ADMIN_ID:
            await message.answer("У вас нет прав для выполнения этого действия.")
            return

        db: Session = SessionLocal()

        now = datetime.now()
        first_day_of_month = now.replace(day=1)
        last_day_of_month = (now.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        name_of_month_on_rus = months_russian[now.month]

        IGNORE_WORKERS = ['1195996440', '6468224924']
        employees = db.query(Employee).filter(not_(Employee.telegram_id.in_(IGNORE_WORKERS))).all()

        data = []
        for employee in employees:
            late_attendances = db.query(Attendance).filter(
                Attendance.employee_id == employee.id,
                Attendance.late == True,
                Attendance.date >= first_day_of_month,
                Attendance.date <= last_day_of_month
            ).all()

            late_days = [attendance.date for attendance in late_attendances]
            late_days_str = ', '.join([str(day) for day in late_days])

            data.append({
                "ФИО": employee.fio,
                "Телеграмм - ID": employee.telegram_id,
                "Телеграмм Ник": employee.full_name,
                "Количество опозданий": len(late_attendances),
                "Дни опозданий": late_days_str
            })

        df = pd.DataFrame(sorted(data, key=lambda x: x["Количество опозданий"], reverse=True))

        report_file = f"Отчет_по_опозданиям_за_{name_of_month_on_rus}.xlsx"
        df.to_excel(report_file, index=False, engine='openpyxl')

        wb = load_workbook(report_file)
        ws = wb.active
        ws.title = "Отчет"

        ws.insert_rows(1)
        ws.merge_cells('A1:E1')
        ws['A1'] = f'Отчет по опозданиям сотрудников за {name_of_month_on_rus}'
        ws['A1'].font = Font(size=16, bold=True)

        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            if column_letter:
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

        for row in ws.iter_rows(min_row=3, max_col=5):
            try:
                late_count = int(row[3].value)
            except (ValueError, TypeError):
                late_count = 0

            if late_count > 3:
                for cell in row:
                    cell.fill = red_fill
                    cell.border = black_border
            else:
                for cell in row:
                    cell.fill = green_fill
                    cell.border = black_border

        wb.save(report_file)

        report_document = FSInputFile(report_file)

        await message.answer_document(report_document,
                                      caption=f"Отчет по опозданиям сотрудников за {name_of_month_on_rus}")

        db.close()

    @date_router.message(lambda message: message.text == "Отъезд")
    async def departure_command_handler(message: Message, state: FSMContext):
        current_departure_time = datetime.now().time().replace(microsecond=0)
        await state.update_data(departure_time_actual=current_departure_time)

        await message.answer("Выберите тип отъезда:", reply_markup=get_reply_type_keyboard())
        await state.set_state(Form.choosing_departure_type)

    @date_router.callback_query(lambda c: c.data.startswith('type_'))
    async def handle_departure_type(callback_query: CallbackQuery, state: FSMContext):
        departure_type = callback_query.data.split('_')[1].capitalize()
        departure_type_mapping = {
            "Object": "Объект",
            "Personal": "Личный"
        }

        if departure_type not in departure_type_mapping:
            await callback_query.message.answer("Пожалуйста, выберите 'Объект' или 'Личный'.")
            return

        departure_type_russian = departure_type_mapping[departure_type]
        await state.update_data(departure_type=departure_type_russian)
        await callback_query.message.answer("Выберите, у кого отпрашиваетесь:", reply_markup=get_supervisor_keyboard())
        await state.set_state(Form.choosing_supervisor)
        await callback_query.answer()

    @date_router.callback_query(lambda c: c.data.startswith('supervisor_'))
    async def handle_supervisor(callback_query: CallbackQuery, state: FSMContext):
        try:
            supervisor_index = int(callback_query.data.split('_', 1)[1])
            supervisors = ["Кексель Кристина", "Тайбупенова Шолпан"]

            if supervisor_index >= len(supervisors):
                await callback_query.message.answer("Пожалуйста, выберите одного из руководителей.")
                return

            supervisor = supervisors[supervisor_index]
            await state.update_data(supervisor=supervisor)
            await callback_query.message.answer("Пожалуйста, напишите причину вашего отсутствия.")
            await state.set_state(Form.waiting_for_reason)
            await callback_query.answer()
        except Exception as e:
            print(f"Error in handle_supervisor: {e}")
            await callback_query.message.answer("Произошла ошибка при обработке вашего выбора.")

    @date_router.message(Form.waiting_for_reason)
    async def handle_absence_reason(message: Message, state: FSMContext):
        reason = message.text
        data = await state.get_data()
        departure_type = data.get("departure_type")
        supervisor = data.get("supervisor")
        departure_time_actual = data.get("departure_time_actual")
        await state.update_data(departure_reason=reason)

        response_text = (
            f"<b>📩 Отчёт об отъезде</b>\n\n"
            f"<b>Тип отъезда:</b> <i>{departure_type}</i>\n"
            f"<b>Время отъезда:</b> <i>{departure_time_actual}</i>\n"
            f"<b>Руководитель:</b> <i>{supervisor}</i>\n"
            f"<b>Причина:</b>\n"
            f"{reason}"
        )
        keyboard = get_return_keyboard()
        await message.answer(response_text, reply_markup=keyboard)

    @date_router.callback_query(lambda c: c.data == 'return')
    async def return_data(callback_query: CallbackQuery, state: FSMContext):
        db: Session = SessionLocal()
        user = db.query(Employee).filter(
            Employee.telegram_id == callback_query.from_user.id
        ).first()
        existing_attendance = db.query(Attendance).filter(
            Attendance.date == date.today(),
            Attendance.employee_id == user.id,
        ).first()


        if existing_attendance:
            data = await state.get_data()

            departure_type = data.get("departure_type")
            departure_reason = data.get("departure_reason")
            supervisor = data.get("supervisor")
            departure_time_actual = data.get("departure_time_actual")
            return_time = datetime.now().time()

            # Update the fields in the existing attendance record
            existing_attendance.departure_type = departure_type
            existing_attendance.departure_reason = departure_reason
            existing_attendance.supervisor = supervisor
            existing_attendance.departure_time_actual = departure_time_actual.replace(microsecond=0)
            existing_attendance.return_time = return_time.replace(microsecond=0)
            db.commit()

            await callback_query.message.answer("Ваше возвращение зафиксировано. Спасибо!")
        else:
            data = await state.get_data()

            attendance = Attendance(
                employee_id=user.id,
                departure_type=data.get("departure_type"),
                departure_reason=data.get("departure_reason"),
                supervisor=data.get("supervisor"),
                departure_time_actual=data.get("departure_time_actual").replace(microsecond=0),
                return_time=datetime.now().time().replace(microsecond=0),
                check=True,
                date=date.today(),
            )
            db.add(attendance)
            db.commit()
            await callback_query.message.answer("Ваше возвращение зафиксировано. Спасибо!")

        await state.clear()
        db.close()


    @date_router.message()
    async def handle_all_messages(message: Message, bot: Bot) -> None:
        await universal_message_handler(message, bot)

async def send_message_to_all_employees(bot: Bot, message_text: str, user_id: int) -> None:
    db: Session = SessionLocal()
    # SEND_MSG = ['1195996440', '6468224924', '1205183489', '1729063947', '595551594', '1600592877']
    # if str(user_id) not in SEND_MSG:
    #     return
    employees = db.query(Employee).all()
    for employee in employees:
        try:
            PRODUCT = os.getenv('PRODUCT')

            if int(PRODUCT):
                await bot.send_message(employee.telegram_id, message_text)
            else:
                print(message_text)
        except Exception as e:
            print(f"Не удалось отправить сообщение сотруднику {employee.telegram_id}: {e}")
    db.close()


async def universal_message_handler(message: Message, bot: Bot) -> None:
    telegram_id = message.from_user.id
    db: Session = SessionLocal()
    employee = db.query(Employee).filter(Employee.telegram_id == telegram_id).first()
    db.close()
    employee_name = employee.fio if employee and employee.fio else message.from_user.full_name
    message_text = (
        f"<b>📩 Новое сообщение</b>\n\n"
        f"<b>От:</b> <i>{employee_name}</i>\n"
        f"<b>Сообщение:</b>\n"
        f"{message.text}"
    )

    await send_message_to_all_employees(bot, message_text, message.from_user.id)